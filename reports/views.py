from collections import Counter
from datetime import datetime

from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from django.shortcuts import render
from django.utils import timezone
from django.http import HttpResponse

from accounts.models import Employee, Oddeleni
from timetracking.models import WorkdaySummary
from leaves.models import TypDovolene, ZadostODovolenou, ZustatekDovolene
from .services import NEPRITOMEN, PRITOMEN, stavy_zamestnancu

DELKA_VYHLEDAVACIHO_DOTAZU = 2
LIMIT_VYSLEDKU_VYHLEDAVANI = 50


def je_admin_nebo_vedouci(user):
    return user.is_staff or hasattr(user, "employee")


@login_required
def prehled_tymu(request):
    """Vedoucí vidí přehled svého týmu za aktuální měsíc."""
    employee = request.user.employee
    dnes = timezone.localdate()
    rok = int(request.GET.get("rok", dnes.year))
    mesic = int(request.GET.get("mesic", dnes.month))

    # Zaměstnanci v oddělení vedoucího
    if request.user.is_staff:
        podrizeni = Employee.objects.filter(aktivni=True)
    else:
        oddeleni = employee.oddeleni
        if oddeleni.vedouci == employee:
            podrizeni = oddeleni.zamestnanci.filter(aktivni=True)
        else:
            podrizeni = Employee.objects.none()

    data = []
    for podr in podrizeni.select_related("user", "typ_uvazku"):
        souhrny = WorkdaySummary.objects.filter(
            employee=podr, datum__year=rok, datum__month=mesic
        )
        celkem = sum(s.odpracovane_minuty for s in souhrny)
        prescos = sum(s.prescos_minuty for s in souhrny)
        data.append({
            "employee": podr,
            "odpracovano_h": celkem // 60,
            "odpracovano_m": celkem % 60,
            "prescos_h": prescos // 60,
            "prescos_m": prescos % 60,
        })

    return render(request, "reports/prehled_tymu.html", {
        "data": data, "rok": rok, "mesic": mesic,
    })


@login_required
def export_xlsx(request):
    """Export měsíčního výkazu do Excelu (openpyxl)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from calendar import monthrange

    employee = request.user.employee
    dnes = timezone.localdate()
    rok = int(request.GET.get("rok", dnes.year))
    mesic = int(request.GET.get("mesic", dnes.month))

    souhrny = {
        s.datum: s
        for s in WorkdaySummary.objects.filter(
            employee=employee, datum__year=rok, datum__month=mesic
        )
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Výkaz {mesic:02d}/{rok}"

    # Záhlaví
    hlavicka = ["Datum", "Den", "Odpracováno", "Přesčas", "Svátek/Víkend"]
    for col, text in enumerate(hlavicka, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4A90E2")

    dny_v_mesici = monthrange(rok, mesic)[1]
    from datetime import date
    nazvy_dnu = ["Po", "Út", "St", "Čt", "Pá", "So", "Ne"]

    for den in range(1, dny_v_mesici + 1):
        datum = date(rok, mesic, den)
        souhrn = souhrny.get(datum)
        row = den + 1

        ws.cell(row=row, column=1, value=datum.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=2, value=nazvy_dnu[datum.weekday()])

        if souhrn:
            odpr = f"{souhrn.odpracovane_minuty // 60}h {souhrn.odpracovane_minuty % 60}min"
            prescos = f"{souhrn.prescos_minuty // 60}h {souhrn.prescos_minuty % 60}min"
            poznamka = []
            if souhrn.je_svatek:
                poznamka.append("Svátek")
            if souhrn.je_vikend:
                poznamka.append("Víkend")
            ws.cell(row=row, column=3, value=odpr)
            ws.cell(row=row, column=4, value=prescos)
            ws.cell(row=row, column=5, value=", ".join(poznamka))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="vykaz_{rok}_{mesic:02d}_{employee.osobni_cislo}.xlsx"'
    )
    wb.save(response)
    return response


def _parsuj_datum(request):
    datum_str = request.GET.get("datum")
    if datum_str:
        try:
            return datetime.strptime(datum_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    return timezone.localdate()


@login_required
def prehled_pritomnosti(request):
    """
    Denní přehled přítomnosti: admin vidí vše, vedoucí oddělení vidí své
    oddělení, ostatní zaměstnanci vidí celý svůj odbor (napříč odděleními).
    """
    datum = _parsuj_datum(request)
    ma_pristup = je_admin_nebo_vedouci(request.user)

    if request.user.is_staff:
        zamestnanci = Employee.objects.filter(aktivni=True)
    elif hasattr(request.user, "employee"):
        employee = request.user.employee
        oddeleni = employee.oddeleni
        if oddeleni.vedouci == employee:
            zamestnanci = oddeleni.zamestnanci.filter(aktivni=True)
        else:
            zamestnanci = Employee.objects.filter(
                oddeleni__odbor=oddeleni.odbor, aktivni=True
            )
    else:
        zamestnanci = Employee.objects.none()

    zamestnanci = list(
        zamestnanci.select_related("user", "oddeleni").order_by(
            "oddeleni", "user__last_name", "user__first_name"
        )
    )
    stavy = stavy_zamestnancu(zamestnanci, datum)
    radky = [{"employee": zam, "stav": stavy[zam.pk]} for zam in zamestnanci]

    poradi_kategorii = [
        (PRITOMEN, "Přítomen"),
        *[(k.value, k.label) for k in TypDovolene.KategoriePrehled],
        (NEPRITOMEN, "Nepřítomen"),
    ]
    pocitadlo = Counter(r["stav"].kod for r in radky)
    pocty = [
        {"popisek": popisek, "pocet": pocitadlo[kod]}
        for kod, popisek in poradi_kategorii
        if pocitadlo[kod]
    ]

    return render(request, "reports/prehled_pritomnosti.html", {
        "datum": datum,
        "radky": radky,
        "pocty": pocty,
        "ma_pristup": ma_pristup,
    })


@login_required
def vyhledat_zamestnance(request):
    """Vyhledání zaměstnance napříč celou firmou (bez omezení na odbor)."""
    dotaz = request.GET.get("q", "").strip()
    vysledky = []
    zkraceno = False

    if len(dotaz) >= DELKA_VYHLEDAVACIHO_DOTAZU:
        zamestnanci = list(
            Employee.objects.filter(
                Q(user__first_name__icontains=dotaz) | Q(user__last_name__icontains=dotaz),
                aktivni=True,
            ).select_related("user", "oddeleni")[: LIMIT_VYSLEDKU_VYHLEDAVANI + 1]
        )
        zkraceno = len(zamestnanci) > LIMIT_VYSLEDKU_VYHLEDAVANI
        zamestnanci = zamestnanci[:LIMIT_VYSLEDKU_VYHLEDAVANI]

        stavy = stavy_zamestnancu(zamestnanci, timezone.localdate())
        vysledky = [{"employee": zam, "stav": stavy[zam.pk]} for zam in zamestnanci]

    return render(request, "reports/vyhledat_zamestnance.html", {
        "dotaz": dotaz,
        "vysledky": vysledky,
        "zkraceno": zkraceno,
        "min_delka_dotazu": DELKA_VYHLEDAVACIHO_DOTAZU,
    })


@login_required
def reports_urls(request):
    return render(request, "reports/index.html")
